# -*- coding: utf-8 -*-

import time
import base64
import urllib
import json

from odoo import fields, models, api, _
from odoo.exceptions import Warning as UserError
from datetime import datetime

from io import BytesIO
from . import utils
from openpyxl.styles import Font, PatternFill

class FacturacionMoveReportWizard(models.TransientModel):
    _name = 'facturacion.move.report.wizard'
    _description = 'Reportes'

    company_id = fields.Many2one('res.company', string='Company', readonly=True, default=lambda self: self.env.user.company_id)

    tipo = fields.Selection([
            ("facturacion","Facturación"),
            ("rubros","Rubros"),
            ("cuenta_x_cobrar_empresas","Cuentas por cobrar empresas"),
            ("notas_credito","Notas de crédito emitidas"),
        ], string=_("Tipo"), default="facturacion", tracking=True)

    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date', required=True)

    xls_export_file = fields.Binary('Excel File')
    xls_export_filename = fields.Char('Excel Filename', size=80, readonly=True)

    def get_invoices(self, invoices):

        data = {}
        i = 0
        estados = {
            'reversed': 'Reversado',
            'paid': 'Pagado',
            'not_paid': 'Sin pagar'
        }
        for invoice in invoices:
            line = invoice.line_ids.filtered(lambda l: l.credit > 0)
            account_id = line[0].account_id.name
            if invoice.tax_totals:
                total = invoice.tax_totals['amount_total']
                subtotal = invoice.tax_totals['amount_untaxed']
            else:
                total = 0
                subtotal = 0
            forma_pago = ""
            date = ""
            if invoice.invoice_payments_widget:
                valor = sum(d['amount'] for d in invoice.invoice_payments_widget['content'])
                for d in invoice.invoice_payments_widget['content']:
                    forma_pago = d['payment_method_name']
                    date = d['date'].strftime("%d/%m/%Y")
            else:
                valor = 0
            if invoice.payment_state in estados:
                estado = estados[invoice.payment_state]
            else:
                estado = 'NN'
            datos = [
                    i,
                    invoice.name,
                    invoice.partner_id.name,
                    subtotal,
                    invoice.subtotal_iva,
                    total,
                    valor,
                    invoice.amount_residual,
                    date,
                    forma_pago,
                    estado
                ]
            if account_id in data:
                data[account_id].append(datos)
            else:
                data.update({ account_id: [datos] })
            i = i + 1
        columnas = []
        i = 0
        cuentas = []
        for account, account_values in data.items():
            columnas.append([account,'','','','','','','','','',''])
            for values in account_values:
                columnas.append(values)
            columnas.append([])
            i += 1
            cuentas.append(i)
        return columnas, cuentas



    def get_rubros(self, invoices):
        # INGRESOS POR RECAUDACIÓN EN VENTAS DE BIENES Y SERVICIOS (RUBROS)
        # detalle por rubros, pvp, cantidad, subtotal, iva, total

        data = {}
        for invoice in invoices:
            lines = invoice.invoice_line_ids
            for line in lines:
                pvp = line.price_unit
                cantidad = line.quantity
                subtotal = line.price_subtotal
                iva = line.tax_base_amount
                total = line.price_total

                account = line.account_id.name
                product = line.product_id.name
                iva = 0
                datos_inv = {
                    'pvp': pvp,
                    'cantidad': cantidad,
                    'subtotal': subtotal,
                    'iva': iva,
                    'total': total
                }
                if account in data:
                    if product in data[account]:
                        data[account][product]['pvp'] += pvp
                        data[account][product]['cantidad'] += cantidad
                        data[account][product]['subtotal'] += subtotal
                        data[account][product]['iva'] += iva
                        data[account][product]['total'] += total
                    else:
                        data[account].update({product: datos_inv})
                else:
                    data.update({
                        account: {
                            product: datos_inv
                        }
                    })
        columnas = []
        cuentas = []
        i = 0
        for account, account_values in data.items():
            columnas.append([account,'','','','',''])
            for product, values in account_values.items():
                columnas.append([product]+list(values.values()))
            columnas.append([])
            i += 1
            cuentas.append(i)
        return columnas, cuentas

    def get_pending_invoices(self, invoices):

        data = {}
        i = 0
        estados = {
            'reversed': 'Reversado',
            'paid': 'Pagado',
            'not_paid': 'Por cobrar'
        }
        data = []
        for invoice in invoices.filtered(lambda a: a.state not in ['paid', 'draft'] and a.payment_state in ['not_paid']):
            line = invoice.line_ids.filtered(lambda l: l.credit > 0)
            account_id = line[0].account_id.name
            if invoice.tax_totals:
                total = invoice.tax_totals['amount_total']
                subtotal = invoice.tax_totals['amount_untaxed']
            else:
                total = 0
                subtotal = 0
            forma_pago = ""
            if invoice.invoice_payments_widget:
                valor = sum(d['amount'] for d in invoice.invoice_payments_widget['content'])
                for d in invoice.invoice_payments_widget['content']:
                    forma_pago = d['payment_method_name']
            else:
                valor = 0
            if invoice.payment_state in estados:
                estado = estados[invoice.payment_state]
            else:
                estado = 'NN'
            data.append([
                    i,
                    invoice.name,
                    invoice.invoice_date,
                    invoice.partner_id.name,
                    line[0].product_id.tipo_documento_id.name,
                    line[0].product_id.name,
                    line[0].account_id.code,
                    total,
                    estado
                ])
            i = i + 1
        return data

    def get_refunds(self, invoices):

        data = {}
        i = 0
        estados = {
            'reversed': 'Reversado',
            'paid': 'Pagado',
            'not_paid': 'Por cobrar'
        }
        data = []
        for invoice in invoices.filtered(lambda a: a.state not in ['draft']):
            line = invoice.line_ids.filtered(lambda l: l.credit > 0)
            account_id = line[0].account_id.name
            if invoice.tax_totals:
                total = invoice.tax_totals['amount_total']
                subtotal = invoice.tax_totals['amount_untaxed']
            else:
                total = 0
                subtotal = 0
            if invoice.payment_state in estados:
                estado = estados[invoice.payment_state]
            else:
                estado = 'NN'
            data.append([
                    i,
                    invoice.name,
                    invoice.partner_id.name,
                    invoice.invoice_date,
                    total,
                    invoice.invoice_origin,
                    invoice.reversed_entry_id.invoice_date,
                    invoice.comprobante_id.fecha_autorizacion,
                    invoice.comprobante_id.numero_autorizacion,
                    estado,
                    invoice.create_uid.name
                ])
            i = i + 1
        return data

    def export_xls(self):
        data = {}
        data['ids'] = self.env.context.get('active_ids', [])
        data['model'] = self.env.context.get('active_model', 'ir.ui.menu')
        data['form'] = self.read(['date_from', 'date_to'])[0]
        date_from = data['form'].get('date_from', time.strftime('%Y-%m-%d')) or ''
        date_to = data['form'].get('date_to', time.strftime('%Y-%m-%d'))

        dmn = [
            ('invoice_date', '>=', date_from),
            ('invoice_date', '<=', date_to),
            ('state', '=', 'posted'),
            ('move_type', '=', 'out_invoice'),
            ]

        if not data['form'].get('date_from'):
            dmn = [
                ('invoice_date', '<=', data['form'].get('date_to')),
                ('state', '=', 'posted'),
                ('move_type', '=', 'out_invoice'),
                ]
        if self.tipo == 'notas_credito':
            dmn = [
                ('invoice_date', '<=', data['form'].get('date_to')),
                ('state', '=', 'posted'),
                ('move_type', '=', 'out_refund'),
                ]

        if date_from:
            subtitle = ['', '', 'Del: ' + date_from.strftime('%Y-%m-%d') + ' Al: ' + date_to.strftime('%Y-%m-%d')]
            export_file = "reporte_" + self.tipo + "_" + date_from.strftime("%d%m%Y") + "_" + date_to.strftime("%d%m%Y") + ".xlsx"
        else:
            subtitle = ['', '', 'Al: ' + date_to.strftime('%Y-%m-%d')]
            export_file = "reporte_" + self.tipo + "_" + date_to.strftime("%d%m%Y") + ".xlsx"

        wb =  utils.prepare_workbook()

        head1 = ['ARMADA DEL ECUADOR']
        company = [self.company_id.name.upper()]
        logo = self.company_id.logo
        tamaño_encabezados = 0
        if self.tipo == 'facturacion':
            title = ['REPORTE FACTURACIÓN Y RECAUDACIÓN POR RUBROS']
            headers = ["", "Secuencial", "Razón social", "Subtotal", "IVA", "TOTAL", "VALOR", "SALDO", "REFERENCIA", "TIPO", "ESTADO"]
            tamaño_encabezados = len(headers)
        elif self.tipo == 'rubros':
            title = ['INGRESOS POR RECAUDACIÓN EN VENTAS DE BIENES Y SERVICIOS (RUBROS)']
            headers = ['Detalle por rubros', 'PVP', 'Cantidad', 'Subtotal', 'IVA', 'Total']
            tamaño_encabezados = len(headers)
        elif self.tipo == 'cuenta_x_cobrar_empresas':
            title = ['CUENTAS POR COBRAR EMPRESAS']
            headers = ['No.', 'FACTURA No.', 'FECHA DE EMISIÓN', 'RAZÓN SOCIAL', 'TIPO DE SERVICIO', 'DETALLE', 'PARTIDA', 'SUBTOTAL', 'ESTADO']
            tamaño_encabezados = len(headers)
        elif self.tipo == 'notas_credito':
            title = ['NOTAS CRÉDITO EMITIDAS']
            headers = ['No.', 'N/C No.', 'NOMBRE/RAZÓN SOCIAL', 'FECHA DE EMISIÓN', 'VALOR', 'FACTURA', 'FECHA FACTURA', 'FECHA AUTORIZACIÓN N/C SRI', 'NÚMERO AUTORIZACIÓN N/C SRI', 'ESTADO', 'GENERADO POR']
            tamaño_encabezados = len(headers)
        else:
            return
        ws =  utils.prepare_worksheet(company, wb, title, subtitle, tamaño_encabezados, head1, logo)

        ws.append([])

        invoices = self.env["account.move"].search(dmn)

        cuentas = False
        if self.tipo == 'facturacion':
            preheaders = ["No", "FACTURACIÓN", "", "", "", "", "RECAUDACIÓN", "", "", "", "ESTADO"]
            ws.append(preheaders)
            utils.header(ws, ws.max_row, ws.max_row)
            ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
            ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=10)
            #headers = ["", "Secuencial", "Razón social", "Subtotal", "IVA", "TOTAL", "VALOR", "SALDO", "Referencia", "Tipo", ""]
            ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
            ws.merge_cells(start_row=6, start_column=11, end_row=7, end_column=11)
            ws.append(headers)
            utils.header(ws, ws.max_row, ws.max_row)
            first_row = ws._current_row
            data, cuentas = self.get_invoices(invoices)
            celdas = ['D','E','F','G','H']
        elif self.tipo == 'rubros':
            #headers = ['Detalle por rubros', 'PVP', 'Cantidad', 'Subtotal', 'IVA', 'Total']
            ws.append(headers)
            utils.header(ws, ws.max_row, ws.max_row)
            first_row = ws._current_row
            data, cuentas = self.get_rubros(invoices)
            celdas = ['B','C','D','E','F']
        elif self.tipo == 'cuenta_x_cobrar_empresas':
            #headers = ['No.', 'FACTURA No.', 'FECHA DE EMISIÓN', 'RAZÓN SOCIAL', 'TIPO DE SERVICIO', 'DETALLE', 'PARTIDA', 'SUBTOTAL', 'ESTADO']
            ws.append(headers)
            utils.header(ws, ws.max_row, ws.max_row)
            first_row = ws._current_row
            data = self.get_pending_invoices(invoices)
            celdas = ['H']
        elif self.tipo == 'notas_credito':
            ws.append(headers)
            utils.header(ws, ws.max_row, ws.max_row)
            first_row = ws._current_row
            data = self.get_refunds(invoices)
            celdas = ['H']
        else:
            return

        for values in data:
            ws.append(values)

        max_row = ws.max_row

        if cuentas:
            print(cuentas)
            for cuenta in cuentas:
                print(cuenta+first_row)
                for cell in ws[cuenta+first_row]:
                    print(cell)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")
                ws.insert_rows(cuenta+first_row, 1)

        for celda in celdas:
            min = celda + str(first_row)
            total = celda + str(max_row)
            total1 = celda + str(max_row+1)
            ws[total1] = '=SUM('+ min + ':'+total+')'

        total = 'A'+str(max_row+1)
        ws[total] = 'TOTAL'

        utils.header(ws, ws.max_row, ws.max_row)
        ws.append([])
        ws.append([])
        ws.append(["GENERADO POR","VTO. BUENO","RECIBIDO POR"])
        ws.append([])
        ws.append([])
        ws.append([self.env.user.rango_id.abreviatura , self.env.user.name, self.env.user.rango_id.abreviatura , self.env.user.name,self.env.user.rango_id.abreviatura , self.env.user.name])
        ws.append(["FACTURADOR", "", "JEFE DE ATENUS", "", "", "TESORERÍA"])

        utils.set_size(ws)
        utils.header(ws, 1, 4)

        virtual_workbook = BytesIO()

        wb.save(virtual_workbook)

        self.write({
            'xls_export_file': base64.encodebytes(virtual_workbook.getvalue()),
            'xls_export_filename': export_file,
            })

        file_url = '/web/content?' + urllib.parse.urlencode(dict(
            model = self._name,
            field = 'xls_export_file',
            filename_field = 'xls_export_filename',
            id = self.id,
            download = True,
            filename = self.xls_export_filename,
            ))

        return {
            'type': 'ir.actions.act_url',
            'url': file_url,
            'target': 'new'
        }
