# -*- coding: utf-8 -*-

from openpyxl import Workbook, drawing
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from PIL import Image as PILImage
import base64, io

def prepare_workbook():
    wb = Workbook()
    return wb

def prepare_worksheet(company, wb, title, subtitle, max_column, head1, logo=None):
    ws = wb.active

    im = PILImage.open(io.BytesIO(base64.b64decode(logo)))
    img = drawing.image.Image(im)
    img.anchor = 'A1'
    img.height = 80
    img.width = 80
    ws.add_image(img)

    ws.append(head1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    ws.append(company)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    ws.append(title)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    ws.append(subtitle)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_column)

    return ws

def draw_table(ws, headers_rows):

    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(size=10)

    ft = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=headers_rows):
        for cell in row:
            cell.font = Font(size=10)
            cell.font = ft
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return ws

def header(ws, inicio, fin):

    ft = Font(bold=True)
    for row in ws.iter_rows(min_row=inicio, max_col=ws.max_column, max_row=fin):
        for cell in row:
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")


def fontbold(ws, r):

    ft = Font(bold=True)
    for row in ws.iter_rows(min_row=r, min_col=1, max_col=ws.max_column, max_row=r):
        for cell in row:
            cell.font = Font(size=10, bold=True)

def set_size(ws):

    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(size=10)
    return ws
