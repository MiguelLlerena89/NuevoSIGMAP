from odoo import api, fields, models, _
from odoo.exceptions import RedirectWarning, UserError, ValidationError, AccessError
from ..wizard import utils
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from io import BytesIO
import base64
import urllib
from datetime import date


class SigmapFacturador(models.Model):
    _inherit = "sigmap.facturador"

    abierta = fields.Boolean(default=False)
    name = fields.Char('Nombre')

    def name_get(self):
        return [(rec.id, '%s - %s' % (rec.caja, rec.punto_emision))
            for rec in self]


class FacturacionSesionCaja(models.Model):
    _name = "facturacion.sesion.caja"
    _inherit = ['mail.thread', 'mail.activity.mixin']

    def _default_caja(self):
        user = self.env.user
        caja = self.env['sigmap.facturador'].search([('user_id', '=', user.id)], limit=1)
        return caja

    def _default_company(self):
        return self.env.user.company_id

    company_id = fields.Many2one('res.company', string=_('Institución'), default=_default_company)

    caja_id = fields.Many2one("sigmap.facturador", default=_default_caja)
    establecimiento = fields.Char(related="caja_id.establecimiento", string="Establecimiento")
    punto_emision = fields.Char(related="caja_id.punto_emision", string="Punto emisión")
    user_id = fields.Many2one(related='caja_id.user_id')
    fecha_apertura = fields.Datetime(string=_('Apertura'),
        index=True)
    fecha_cierre = fields.Datetime(string=_('Cierre'), index=True)
    valor_apertura = fields.Float("Valor apertura", default=0,)
    valor_cierre = fields.Float("Valor cierre", compute="_compute_get_valores")
    total_facturado = fields.Float("Total facturado", compute="_compute_get_valores")
    total_nc = fields.Float("Total Notas de crédito", compute="_compute_get_valores")
    payment_ids = fields.One2many("account.payment", "sesion_caja_id", "Recaudaciones", compute="_compute_posted_payments")
    invoice_ids = fields.One2many("account.move", "sesion_caja_id", "Facturas", compute="_compute_posted_invoices")

    xls_caja_reporte_file = fields.Binary('Reporte de caja excel')
    xls_caja_reporte_filename = fields.Char('Nombre reporte de caja excel', size=80, readonly=True)

    xls_facturacion_reporte_file = fields.Binary('Reporte de facturación excel')
    xls_facturacion_reporte_filename = fields.Char('Nombre reporte de facturación excel', size=80, readonly=True)

    @api.depends("caja_id", "fecha_apertura", "fecha_cierre")
    def _compute_posted_invoices(self):
        moves = self.env['account.move'].search(
            domain=[
                ('sesion_caja_id', 'in', self.ids),
                ('state', '!=', 'draft'),
                ('move_type', 'in', self.env['account.move'].get_sale_types())
            ],
        )
        self.invoice_ids = moves

    @api.depends("caja_id", "fecha_apertura", "fecha_cierre")
    def _compute_posted_payments(self):
        moves = self.env['account.payment'].search(
            domain=[
                ('sesion_caja_id', 'in', self.ids),
                ('state', '!=', 'draft'),
                #('move_type', 'in', ['entry'])
            ],
        )
        for move in moves:
            self.write({'payment_ids': (0,0, move.id)})
        else:
            self.write({'payment_ids': False})

    xls_export_file = fields.Binary('Excel File')
    xls_export_filename = fields.Char('Excel Filename', size=80, readonly=True)

    @api.depends("caja_id", "fecha_apertura", "fecha_cierre")
    def _compute_get_valores(self):
        self.valor_cierre = sum(d['amount_total'] for d in self.payment_ids)
        self.total_facturado = sum(d['amount_total'] for d in self.invoice_ids.filtered(lambda a: a.move_type in ['out_invoice'] and a.state not in ['cancel']))
        self.total_nc = sum(d['amount_total'] for d in self.invoice_ids.filtered(lambda a: a.move_type in ['out_refund'] and a.state not in ['cancel']))

    def name_get(self):
        return [(rec.id, '%s - %s' % (rec.caja_id.caja, rec.fecha_apertura))
            for rec in self]

    def get_invoice_report(self, invoices):

        data_invoices = []
        data_refund = []
        estados = {
            'reversed': 'Reversado',
            'paid': 'Pagado',
            'not_paid': 'Sin pagar'
        }
        for invoice in invoices:
            total = 0
            valor_pagado = 0
            if invoice.tax_totals:
                total = invoice.tax_totals['amount_total']
            else:
                continue
            if invoice.invoice_payments_widget:
                pago = {}
                codigo_pago = ''
                valor_pagado = sum(d['amount'] for d in invoice.invoice_payments_widget['content'])
                for d in invoice.invoice_payments_widget['content']:
                    payment = self.env['account.payment'].browse(d['account_payment_id'])
                    codigo_pago = payment.payment_code
            if invoice.move_type == 'out_invoice':
                data_invoices.append([
                    codigo_pago,
                    invoice.name,
                    valor_pagado,
                    total,
                    total - valor_pagado,
                    ''
                ])
            if invoice.move_type == 'out_refund':
                data_refund.append([
                    codigo_pago,
                    valor_pagado,
                    invoice.name,
                    total,
                    ''
                ])
        return data_invoices, data_refund


    def get_invoices(self, invoices):

        data = []
        estados = {
            'reversed': 'Reversado',
            'paid': 'Pagado',
            'not_paid': 'Sin pagar'
        }
        facturas_emitidas = []
        total_facturado = 0
        total_por_cobrar = 0
        total_recaudado = 0
        pago = {}
        for invoice in invoices:
            facturas_emitidas.append(invoice.numero)
            line = invoice.line_ids.filtered(lambda l: l.credit > 0)
            account_id = line[0].account_id.code
            if invoice.tax_totals:
                total = invoice.tax_totals['amount_total']
            else:
                continue
            if invoice.invoice_payments_widget:
                valor_pagado = sum(d['amount'] for d in invoice.invoice_payments_widget['content'])
                for d in invoice.invoice_payments_widget['content']:
                    if d['payment_method_name'] in pago:
                        pago[d['payment_method_name']] += d['amount']
                    else:
                        pago[d['payment_method_name']] = d['amount']
                    forma_pago = d['payment_method_name']
            else:
                valor_pagado = 0
                forma_pago = ""
            data.append([
                line[0].product_id.tipo_documento_id.name,
                account_id,
                line[0].name,
                total,
                valor_pagado,
                '',
                forma_pago,
                invoice.amount_residual,
            ])
            total_facturado += total
            total_por_cobrar += invoice.amount_residual
            total_recaudado += valor_pagado
        totales = {
            'total_facturado': total_facturado,
            'total_por_cobrar': total_por_cobrar,
            'total_recaudado': total_recaudado,
            'total_deposito': 0,
            'total_transferencia': 0,
            'total_cheque': 0,
            'total_appmovil': 0
        }
        if 'Depósito' in pago:
            totales['total_deposito'] = pago['Deposito']
        if 'Transferencia' in pago:
            totales['total_transferencia']= pago['Transferencia']
        if 'Cheque' in pago:
            totales['total_cheque'] = pago['Cheque']
        if 'App Móvil' in pago:
            totales['total_appmovil'] = pago['App Móvil']

        return data, totales

    def print_xlsx_caja_report(self):
        self.ensure_one()

        wb =  utils.prepare_workbook()

        head1 = ['ARMADA DEL ECUADOR']
        company = [self.company_id.name.upper()]
        reparto = [self.company_id.name.upper()]
        logo = self.company_id.logo
        title = ['REPORTE DIARIO DE CAJA ELECTRÓNICA']
        subtitle = [self.company_id.city]
        export_file = "reporte_diario_" + self.caja_id.name + self.fecha_cierre.strftime("%d%m%Y") + ".xlsx"
        len = 7
        ws =  utils.prepare_worksheet(company, wb, title, subtitle, len, head1, logo)

        ws.append([])
        fecha = ['', '', '', '', '', ' Fecha: ' + self.fecha_cierre.strftime('%Y-%m-%d')]
        ws.append(fecha)
        detalle_caja = ['ESTABLECIMIENTO/CAJA No.', self.caja_id.sri_establecimiento_id.code]
        ws.append(detalle_caja)


        invoices = self.invoice_ids
        facts_emitidas = [ inv.numero for inv in self.invoice_ids.filtered(lambda a: a.move_type in ['out_invoice'] and a.state in ['posted'])]
        facts_anuladas = [ inv.numero for inv in self.invoice_ids.filtered(lambda a: a.move_type in ['out_invoice'] and a.state in ['cancel'])]
        notas_credito = [ inv.numero for inv in self.invoice_ids.filtered(lambda a: a.move_type in ['out_refund'] and a.state in ['posted'])]

        listado_facts = ["FACTURAS EMITIDAS #"] + [' , '.join(facts_emitidas)]
        ws.append(listado_facts)
        ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row+1, end_column=len+1)
        ws.append([])
        listado_facts_anuladas = ["FACTURAS ANULADAS #", ' , '.join(facts_anuladas)]
        ws.append(listado_facts_anuladas)
        ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row+1, end_column=len+1)
        ws.append([])
        preheaders = ["CONCEPTO", "PARTIDAS", "DETALLE", "TOTAL FACTURADO", "VALOR RECAUDADO", "CXC (Instituciones públicas)", "FORMA DE PAGO", "SALDOS"]
        ws.append(preheaders)
        headers = ["FACTURAS"]
        ws.append(headers)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len+1)
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")
        first_row = ws._current_row
        data, totales = self.get_invoices(invoices)
        data_nc, totales_nc = self.get_invoices(notas_credito)
        celdas = ['D','E','F','H']

        for values in data:
            ws.append(values)

        max_row = ws.max_row

        for celda in celdas:
            min = celda + str(first_row)
            total = celda + str(max_row)
            total1 = celda + str(max_row+1)
            ws[total1] = '=SUM('+ min + ':'+total+')'

        total = 'A'+str(max_row+1)
        ws[total] = 'SUBTOTAL'
        for row in ws.iter_rows(min_row=ws.max_row, max_col=len+1, max_row=ws.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")

        ws.append([])
        ws.append(["RESUMEN DE FACTURACIÓN"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len+1)
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        current_cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")
        ws.append(["FACTURADO", "", "", "", "", "", "", totales['total_facturado']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)
        ws.append(["VALOR POR COBRAR", "", "", "", "", "", "",  totales['total_por_cobrar']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)
        ws.append(["TOTAL RECAUDACIÓN", "", "", "", "", "", "", totales['total_recaudado']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append([])
        ws.append(["RESUMEN RECAUDACIÓN"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len+1)
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        current_cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")

        ws.append(["DEPÓSITOS/CÓDIGOS", "", "", "", "", "", "", totales['total_deposito']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append(["TRANSFERENCIAS", "", "", "", "", "", "",  totales['total_transferencia']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append(["CHEQUE", "", "", "", "", "", "",  totales['total_cheque']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append(["Aplicación MÓVIL", "", "", "", "", "", "",  totales['total_appmovil']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append(["LIQUIDACIÓN ACREEDOR", "", "", "", "", "", "", ""])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append(["TOTAL", "", "", "", "", "", "", totales['total_recaudado']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append([])
        ws.append(["NOTAS DE CRÉDITO EMITIDAS", "", "", "", "", "", "", totales_nc['total_facturado']])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column-1)

        ws.append([])
        current_cell = ws.cell(row=ws.max_row+1, column=1)
        current_cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")

        ws.append(["", "", "GENERADO POR", "", "", "", "", "VTO. BUENO"])
        ws.append(["", "", self.env.user.rango_id.abreviatura, self.env.user.name, "", "", self.env.user.rango_id.abreviatura, self.env.user.name])
        ws.append(["", "", "FACTURACIÓN", "", "", "", "", "JEFE DE ATENUS"])
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "RECIBIDO POR DPTO FINANCIERO"])
        ws.append(["", "",  "", "", "", "",self.env.user.rango_id.abreviatura, self.env.user.name])
        ws.append([])
        ws.append(["Impreso por:", self.env.user.name, "", "Fecha", date.today().strftime("%d/%m/%Y")])

        utils.set_size(ws)
        utils.header(ws, 1, 4)

        virtual_workbook = BytesIO()

        wb.save(virtual_workbook)

        self.write({
            'xls_caja_reporte_file': base64.encodebytes(virtual_workbook.getvalue()),
            'xls_caja_reporte_filename': export_file,
            })

        file_url = '/web/content?' + urllib.parse.urlencode(dict(
            model = self._name,
            field = 'xls_caja_reporte_file',
            filename_field = 'xls_caja_reporte_filename',
            id = self.id,
            download = True,
            filename = self.xls_caja_reporte_filename,
            ))

        return {
            'type': 'ir.actions.act_url',
            'url': file_url,
            'target': 'new'
        }


    def print_xlsx_facturacion(self):
        self.ensure_one()

        wb =  utils.prepare_workbook()

        head1 = ['ARMADA DEL ECUADOR']
        company = [self.company_id.name.upper()]
        logo = self.company_id.logo
        title = ['REPORTE FACTURACIÓN']
        subtitle = [self.company_id.city]
        export_file = "reporte_facturacion_" + self.caja_id.user_id.name + '_' + self.fecha_cierre.strftime("%d%m%Y") + ".xlsx"
        preheaders = ["Nro. PAPELETA/COD RECAUDACION", "Nro. FACTURA", "VALOR DEPOSITADO", "VALOR FACTURADO", "CTAS. POR COBRAR", "NOVEDADES"]
        size = len(preheaders)
        ws =  utils.prepare_worksheet(company, wb, title, subtitle, size, head1, logo)

        ws.append([])
        fecha = ['', '', '', '', 'Fecha:', self.fecha_cierre.strftime('%Y-%m-%d')]
        ws.append(fecha)

        invoices = self.invoice_ids

        ws.append([])
        ws.append(preheaders)
        utils.header(ws, ws.max_row, ws.max_row)
        first_row = ws._current_row
        invoices, refunds = self.get_invoice_report(invoices)
        celdas = ['C','D','E']

        for values in invoices:
            ws.append(values)

        max_row = ws.max_row

        for celda in celdas:
            min = celda + str(first_row)
            total = celda + str(max_row)
            total1 = celda + str(max_row+1)
            ws[total1] = '=SUM('+ min + ':'+total+')'

        total = 'A'+str(max_row+1)
        ws[total] = 'SUMAN'

        ws.append([])
        ws.append(['NOTAS DE CRÉDITO EMITIDAS / FACTURAS ANULADAS'])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=size)
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        current_cell.fill = PatternFill(fgColor="D0D1D9", fill_type="solid")
        ws.append(["Nro. PAPELETA DEPOSITO", "VALOR DEPOSITADO", "Nro. FACTURAS", "VALOR FACTURADO", "JUSTIFICACIÓN", ""])
        last_row = ws.max_row

        for values in refunds:
            ws.append(values)

        max_row = ws.max_row

        celdas = ['B','D']
        for celda in celdas:
            min = celda + str(last_row)
            total = celda + str(max_row)
            total1 = celda + str(max_row+1)
            ws[total1] = '=SUM('+ min + ':'+total+')'

        total = 'A'+str(max_row+1)
        ws[total] = 'SUMAN'

        ws.append([])
        ws.append([])
        ws.append([])
        ws.append([])
        ws.append(["", "FACTURADOR", "", "", "", "JEFE DE ATENUS"])
        ws.append([])
        ws.append(["FECHA Y HORA DE CIERRE DE CAJA", self.fecha_cierre])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.font = Font(size=10, bold=True)
        ws.append(["MÓDULO DE FACTURACIÓN", self.env.user.rango_id.abreviatura, self.env.user.name])
        current_cell = ws.cell(row=ws.max_row, column=1)
        current_cell.font = Font(size=10, bold=True)

        utils.set_size(ws)
        utils.header(ws, 1, 4)

        virtual_workbook = BytesIO()

        wb.save(virtual_workbook)

        self.write({
            'xls_facturacion_reporte_file': base64.encodebytes(virtual_workbook.getvalue()),
            'xls_facturacion_reporte_filename': export_file,
            })

        file_url = '/web/content?' + urllib.parse.urlencode(dict(
            model = self._name,
            field = 'xls_facturacion_reporte_file',
            filename_field = 'xls_facturacion_reporte_filename',
            id = self.id,
            download = True,
            filename = self.xls_facturacion_reporte_filename,
            ))

        return {
            'type': 'ir.actions.act_url',
            'url': file_url,
            'target': 'new'
        }
